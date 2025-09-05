# encoding: utf-8
'''
@author: lingshu
@file: read_excel.py
@time: 2019/6/21 17:03
@desc: 读取excel
'''
import xlrd
import os
from utils import get_file_path
def read_excel():
    file_path = get_file_path.get_root_path()+'testdata\\testdata.xlsx'
    # 文件位置
    excel_file=xlrd.open_workbook(file_path)

    # 获取sheet内容【1.根据sheet索引2.根据sheet名称】
    # sheet=ExcelFile.sheet_by_index(1)
    sheet = excel_file.sheet_by_name('Sheet1')
    # 打印sheet的名称，行数，列数
    print(sheet.name)
    print(sheet.nrows)
    print(sheet.ncols)

    # 获取整行或者整列的值
    rows = sheet.row_values(1)
    cols = sheet.col_values(1)
    print(rows)
    print(cols)

    #获取单元格内容
    print("第二行第一列的值为： %s",sheet.cell(1,0))

    # 打印单元格内容格式
    print("单元格内容格式为： %s",sheet.cell(0,0).ctype)
"""
def get_xls():
    cls = []
    print(os.getcwd())
    file_path = os.getcwd() + '\\testdata\\testdata.xlsx'
    # 文件位置
    excel_file = xlrd.open_workbook(file_path)
    sheet = excel_file.sheet_by_name('Sheet1')
    nrows = sheet.nrows
    for i in range(nrows):
        cls.append(sheet.row_values(i))
    return cls
"""
# utils/read_excel.py
import os
from openpyxl import load_workbook

def get_xls():
    cls = []
    print("当前工作目录:", os.getcwd())
    
    # 构建文件路径：项目根目录/testdata/testdata.xlsx
    file_path = os.path.join(os.getcwd(), 'testdata', 'testdata.xlsx')
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel 文件未找到: {file_path}")

    # 使用 openpyxl 读取 .xlsx 文件
    workbook = load_workbook(file_path, data_only=True)  # data_only=True 读取计算后的值，而非公式
    sheet = workbook["Sheet1"]  # 指定 sheet 名

    # 遍历每一行
    for row in sheet.iter_rows(values_only=True):
        cls.append(list(row))  # 将元组转为列表，保持与 xlrd 行为一致

    workbook.close()
    return cls
if __name__ == '__main__':
    print(get_xls())